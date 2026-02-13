"""
QB Diagnostic Report — Excel Export.

Generates a professional Excel workbook (.xlsx) from a diagnostic report dict.
Three sheets: Gap Analysis, Unmapped QB Accounts, Summary.

Uses openpyxl (already in requirements.txt as openpyxl==3.1.5).
"""

from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


# ---------------------------------------------------------------------------
# Color palette
# ---------------------------------------------------------------------------

_HEADER_FILL = PatternFill(start_color="2F3640", end_color="2F3640", fill_type="solid")
_HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_LABEL_FONT = Font(name="Calibri", bold=True, size=11)
_VALUE_FONT = Font(name="Calibri", size=11)
_TITLE_FONT = Font(name="Calibri", bold=True, size=14)

_ROW_FILL_MATCHED = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
_ROW_FILL_CANDIDATES = PatternFill(start_color="FEF9E7", end_color="FEF9E7", fill_type="solid")
_ROW_FILL_UNMATCHED = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")

_GRADE_FILL = {
    "A": PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid"),
    "B": PatternFill(start_color="2E86C1", end_color="2E86C1", fill_type="solid"),
    "C": PatternFill(start_color="E67E22", end_color="E67E22", fill_type="solid"),
    "F": PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid"),
}
_GRADE_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=13)

_THIN_BORDER = Border(
    left=Side(style="thin", color="D5D8DC"),
    right=Side(style="thin", color="D5D8DC"),
    top=Side(style="thin", color="D5D8DC"),
    bottom=Side(style="thin", color="D5D8DC"),
)

_CENTER = Alignment(horizontal="center", vertical="center")
_LEFT_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)
_LEFT = Alignment(horizontal="left", vertical="center")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _safe(value: Any, default: str = "") -> str:
    """Convert value to string, replacing None with a default."""
    if value is None:
        return default
    return str(value)


def _apply_header_row(ws: Worksheet, headers: list[str], row: int = 1) -> None:
    """Write a bold header row with dark fill and white text."""
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER
        cell.border = _THIN_BORDER


def _auto_column_widths(
    ws: Worksheet,
    min_width: int = 10,
    max_width: int = 45,
) -> None:
    """Set column widths based on content, clamped to [min_width, max_width]."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                cell_len = len(str(cell.value or ""))
                if cell_len > max_len:
                    max_len = cell_len
            except Exception:
                pass
        adjusted = min(max(max_len + 3, min_width), max_width)
        ws.column_dimensions[col_letter].width = adjusted


def _apply_border_to_range(
    ws: Worksheet,
    min_row: int,
    max_row: int,
    min_col: int,
    max_col: int,
) -> None:
    """Apply thin borders to a rectangular cell range."""
    for row in ws.iter_rows(
        min_row=min_row, max_row=max_row,
        min_col=min_col, max_col=max_col,
    ):
        for cell in row:
            cell.border = _THIN_BORDER


# ---------------------------------------------------------------------------
# Sheet 1: Gap Analysis
# ---------------------------------------------------------------------------

_GAP_HEADERS = [
    "#",
    "Mapping Type",
    "Template Account",
    "Account Type",
    "Sub Type",
    "Default?",
    "Status",
    "Best Match",
    "Match Score",
    "Confidence",
    "Decision",
    "Decision Account",
]


def _build_gap_analysis(wb: Workbook, report: dict) -> None:
    """Build the Gap Analysis sheet."""
    ws = wb.active
    ws.title = "Gap Analysis"

    _apply_header_row(ws, _GAP_HEADERS)

    items: list[dict] = report.get("items", [])

    for idx, item in enumerate(items, start=1):
        row = idx + 1  # row 1 is header

        best: dict | None = item.get("best_match")
        status = _safe(item.get("status"))

        values = [
            idx,
            _safe(item.get("mapping_type")),
            _safe(item.get("template_account_name")),
            _safe(item.get("template_account_type")),
            _safe(item.get("template_account_sub_type")),
            "Yes" if item.get("is_default") else "No",
            status.title(),
            _safe(best.get("qb_account_name") if best else None),
            round(best["score"], 2) if best and best.get("score") is not None else "",
            _safe(best.get("confidence") if best else None).title(),
            _safe(item.get("decision"), default="Pending").replace("_", " ").title(),
            _safe(item.get("decision_account_name")),
        ]

        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.alignment = _LEFT if col_idx > 1 else _CENTER
            cell.font = _VALUE_FONT

        # Row coloring by status
        if status == "matched":
            fill = _ROW_FILL_MATCHED
        elif status == "candidates":
            fill = _ROW_FILL_CANDIDATES
        else:
            fill = _ROW_FILL_UNMATCHED

        for col_idx in range(1, len(_GAP_HEADERS) + 1):
            ws.cell(row=row, column=col_idx).fill = fill

    # Borders
    if items:
        _apply_border_to_range(ws, 1, len(items) + 1, 1, len(_GAP_HEADERS))

    # Auto-filter and freeze
    ws.auto_filter.ref = f"A1:{get_column_letter(len(_GAP_HEADERS))}{len(items) + 1}"
    ws.freeze_panes = "A2"

    _auto_column_widths(ws)


# ---------------------------------------------------------------------------
# Sheet 2: Unmapped QB Accounts
# ---------------------------------------------------------------------------

_UNMAPPED_HEADERS = [
    "#",
    "QB Account Name",
    "Account Type",
    "Sub Type",
    "Suggested POS Type",
]


def _build_unmapped_accounts(wb: Workbook, report: dict) -> None:
    """Build the Unmapped QB Accounts sheet."""
    ws = wb.create_sheet("Unmapped QB Accounts")

    _apply_header_row(ws, _UNMAPPED_HEADERS)

    unmapped: list[dict] = report.get("unmapped_qb_accounts", [])

    for idx, acct in enumerate(unmapped, start=1):
        row = idx + 1
        values = [
            idx,
            _safe(acct.get("qb_account_name")),
            _safe(acct.get("qb_account_type")),
            _safe(acct.get("qb_account_sub_type")),
            _safe(acct.get("suggested_mapping_type")),
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.alignment = _LEFT if col_idx > 1 else _CENTER
            cell.font = _VALUE_FONT

    # Borders
    if unmapped:
        _apply_border_to_range(ws, 1, len(unmapped) + 1, 1, len(_UNMAPPED_HEADERS))

    # Auto-filter and freeze
    last_row = max(len(unmapped) + 1, 1)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(_UNMAPPED_HEADERS))}{last_row}"
    ws.freeze_panes = "A2"

    _auto_column_widths(ws)


# ---------------------------------------------------------------------------
# Sheet 3: Summary
# ---------------------------------------------------------------------------

_GRADE_RECOMMENDATIONS = {
    "A": [
        "All template mappings have been auto-matched to existing QB accounts.",
        "You can apply these mappings immediately — no manual intervention needed.",
        "Run a health check periodically to detect account renames or deletions.",
    ],
    "B": [
        "Most mappings are matched. A few need manual review.",
        "Review the 'Candidates' items in the Gap Analysis sheet and confirm the best match.",
        "Once all decisions are made, apply the mappings to start syncing.",
    ],
    "C": [
        "Partial coverage — several mappings could not be matched automatically.",
        "Review unmatched items and decide whether to create new QB accounts or map to existing ones.",
        "Consider whether this template is the best fit for the client's Chart of Accounts.",
        "You may need to create missing accounts in QuickBooks before applying.",
    ],
    "F": [
        "Low coverage — the template does not match the client's QB setup well.",
        "Consider switching to a different template that better matches the client's industry.",
        "Alternatively, create the missing accounts in QuickBooks first, then re-run the diagnostic.",
        "Do NOT apply mappings in this state — sync will fail for unmapped account types.",
    ],
}


def _build_summary(wb: Workbook, report: dict) -> None:
    """Build the Summary sheet with a row-based layout."""
    ws = wb.create_sheet("Summary")

    # Title
    ws.merge_cells("A1:B1")
    title_cell = ws.cell(row=1, column=1, value="QuickBooks Diagnostic Report")
    title_cell.font = _TITLE_FONT
    title_cell.alignment = Alignment(horizontal="left", vertical="center")

    # Summary rows (start at row 3 for breathing room)
    current_row = 3
    items: list[dict] = report.get("items", [])
    unmapped: list[dict] = report.get("unmapped_qb_accounts", [])

    grade = _safe(report.get("health_grade"), default="?")

    summary_data: list[tuple[str, Any]] = [
        ("Template", f"{_safe(report.get('template_name'))} ({_safe(report.get('template_key'))})"),
        ("Date", _safe(report.get("created_at", ""))[:19].replace("T", " ")),
        ("Health Grade", grade),
        ("", ""),  # blank spacer
        ("Total Template Mappings", report.get("total_template_mappings", 0)),
        ("Matched (auto-suggested)", report.get("matched", 0)),
        ("Candidates (needs review)", report.get("candidates", 0)),
        ("Unmatched", report.get("unmatched", 0)),
        ("Coverage", f"{report.get('coverage_pct', 0)}%"),
        ("", ""),  # blank spacer
        ("QB Accounts in Client", report.get("total_qb_accounts", 0)),
        ("Unmapped QB Accounts", len(unmapped)),
    ]

    for label, value in summary_data:
        label_cell = ws.cell(row=current_row, column=1, value=label)
        value_cell = ws.cell(row=current_row, column=2, value=value)

        if label:
            label_cell.font = _LABEL_FONT
            label_cell.alignment = _LEFT
            value_cell.font = _VALUE_FONT
            value_cell.alignment = _LEFT

        # Special formatting for the grade cell
        if label == "Health Grade":
            value_cell.fill = _GRADE_FILL.get(grade, PatternFill())
            value_cell.font = _GRADE_FONT
            value_cell.alignment = _CENTER

        current_row += 1

    # Blank row before recommendations
    current_row += 1

    # Recommendations section
    rec_header = ws.cell(row=current_row, column=1, value="Recommendations")
    rec_header.font = Font(name="Calibri", bold=True, size=12)
    current_row += 1

    recommendations = _GRADE_RECOMMENDATIONS.get(grade, [])
    for rec in recommendations:
        bullet_cell = ws.cell(row=current_row, column=1, value=f"\u2022  {rec}")
        bullet_cell.font = _VALUE_FONT
        bullet_cell.alignment = _LEFT_WRAP
        ws.merge_cells(
            start_row=current_row, start_column=1,
            end_row=current_row, end_column=2,
        )
        current_row += 1

    # Column widths for summary sheet
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 55


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def generate_diagnostic_excel(report: dict) -> bytes:
    """
    Generate a professional Excel workbook from a diagnostic report.

    Args:
        report: The diagnostic report dict (from DiagnosticService.run_diagnostic
                or DiagnosticService.get_report).

    Returns:
        The .xlsx file contents as bytes, ready to be sent as an HTTP response
        with content-type application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.
    """
    wb = Workbook()

    # Sheet 1: Gap Analysis (uses the default active sheet)
    _build_gap_analysis(wb, report)

    # Sheet 2: Unmapped QB Accounts
    _build_unmapped_accounts(wb, report)

    # Sheet 3: Summary
    _build_summary(wb, report)

    # Write to bytes buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()
